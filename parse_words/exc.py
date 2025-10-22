from openpyxl import Workbook
from openpyxl.styles import Font
from main import parse_word_info
from utils import sleep_rand_time
from data.list_words import words_to_parse

wb = Workbook()
ws = wb.active

# ws1 = wb.create_sheet('MySheet', 0)


ws.title = 'New Title1'

d = {
    'A': 20,
    'B': 20,
    'C': 10,
    'D': 14,
    'E': 20,
    'F': 60,
    'G': 20,
    'H': 20,
    'I': 20,
    'J': 20,
    'K': 20,
    'L': 20,
    'M': 20,
}

for key in d:
    ws.column_dimensions[key].width = d[key]

font_style_title = Font(bold=True, size=12)


rows_title_list = ['word', 'trans', 'trans_type', 'part of speech', 'sense', 'definition', 'example 1', 'example 2', 'example 3', 'example 4', 'example 5', 'example 6', 'example 7']

for column, column_title in enumerate(rows_title_list):
    ws.cell(row=1, column=column+1).value = column_title
    ws.cell(row=1, column=column+1).font = font_style_title


parsed_words = []

len_words_to_parse = len(words_to_parse)
total_time = 0

for counter, word_to_parse in enumerate(words_to_parse):

    parsed_word = None

    try:
        parsed_word = parse_word_info(word_to_parse)
    except Exception as e:
        print(f'Problem with word: {word_to_parse}')
        print(e)

    print(f'{counter + 1} of {len_words_to_parse}', end=' ')

    if len_words_to_parse > 10:
        total_time += sleep_rand_time(min_sec=10, max_sec=25)
    else:
        total_time += sleep_rand_time()

    if parsed_word:
        parsed_words.append(parsed_word)

print(f'total time for {len_words_to_parse} word(s): {round(total_time, 2)}s.', end=' ')

filled_rows = 1

for word_to_excel in parsed_words:

    for row_num, data in enumerate(word_to_excel):
        list_data = list(data.values())

        for column_num, column_data in enumerate(list_data):
            ws.cell(row=filled_rows+row_num+1, column=column_num+1).value = column_data

    filled_rows += len(word_to_excel)

wb.save('check.xlsx')
